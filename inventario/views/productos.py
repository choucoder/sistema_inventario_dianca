from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from ..models import Product, Category, Entrada, InventoryAdjustment


@login_required
def producto_list(request):
    productos = Product.objects.select_related('category').all().order_by('name')
    return render(request, 'productos/list.html', {'productos': productos})


@login_required
def producto_create(request):
    categorias = Category.objects.filter(status='active').order_by('name')

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        unit = request.POST.get('unit', '').strip()
        min_stock = request.POST.get('min_stock', '0').strip()
        category_id = request.POST.get('category', '')
        location = request.POST.get('location', '').strip()

        errors = False

        if not code:
            messages.error(request, 'El codigo del producto es requerido.')
            errors = True
        elif len(code) > 50:
            messages.error(request, 'El codigo no puede exceder 50 caracteres.')
            errors = True

        if code and Product.objects.filter(code__iexact=code).exists():
            messages.error(request, 'Ya existe un producto con ese codigo.')
            errors = True

        if not name:
            messages.error(request, 'El nombre del producto es requerido.')
            errors = True
        elif len(name) > 200:
            messages.error(request, 'El nombre no puede exceder 200 caracteres.')
            errors = True

        if not unit:
            messages.error(request, 'La unidad de medida es requerida.')
            errors = True
        elif len(unit) > 20:
            messages.error(request, 'La unidad no puede exceder 20 caracteres.')
            errors = True

        if not category_id:
            messages.error(request, 'La categoria es requerida.')
            errors = True

        if len(location) > 100:
            messages.error(request, 'La ubicacion no puede exceder 100 caracteres.')
            errors = True

        try:
            min_stock_int = int(min_stock) if min_stock else 0
            if min_stock_int < 0:
                messages.error(request, 'El stock minimo no puede ser negativo.')
                errors = True
        except ValueError:
            messages.error(request, 'El stock minimo debe ser un numero entero.')
            errors = True
            min_stock_int = 0

        if errors:
            return render(request, 'productos/form.html', {
                'categorias': categorias,
                'code': code,
                'name': name,
                'description': description,
                'unit': unit,
                'min_stock': min_stock,
                'category_id': category_id,
                'location': location,
            })

        category = get_object_or_404(Category, pk=category_id)

        Product.objects.create(
            code=code,
            name=name,
            description=description or None,
            unit=unit,
            min_stock=min_stock_int,
            stock_actual=0,
            category=category,
            location=location,
            status='active'
        )
        messages.success(request, f'Producto "{name}" creado exitosamente.')
        return redirect('producto_list')

    return render(request, 'productos/form.html', {
        'categorias': categorias,
        'code': '',
        'name': '',
        'description': '',
        'unit': '',
        'min_stock': '0',
        'category_id': '',
        'location': '',
    })


@login_required
def producto_edit(request, pk):
    producto = get_object_or_404(Product, pk=pk)
    categorias = Category.objects.filter(status='active').order_by('name')

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        unit = request.POST.get('unit', '').strip()
        min_stock = request.POST.get('min_stock', '0').strip()
        category_id = request.POST.get('category', '')
        location = request.POST.get('location', '').strip()
        status = request.POST.get('status', 'active')

        errors = False

        if not code:
            messages.error(request, 'El codigo del producto es requerido.')
            errors = True
        elif len(code) > 50:
            messages.error(request, 'El codigo no puede exceder 50 caracteres.')
            errors = True

        if code and Product.objects.filter(code__iexact=code).exclude(pk=pk).exists():
            messages.error(request, 'Ya existe otro producto con ese codigo.')
            errors = True

        if not name:
            messages.error(request, 'El nombre del producto es requerido.')
            errors = True
        elif len(name) > 200:
            messages.error(request, 'El nombre no puede exceder 200 caracteres.')
            errors = True

        if not unit:
            messages.error(request, 'La unidad de medida es requerida.')
            errors = True
        elif len(unit) > 20:
            messages.error(request, 'La unidad no puede exceder 20 caracteres.')
            errors = True

        if not category_id:
            messages.error(request, 'La categoria es requerida.')
            errors = True

        if len(location) > 100:
            messages.error(request, 'La ubicacion no puede exceder 100 caracteres.')
            errors = True

        try:
            min_stock_int = int(min_stock) if min_stock else 0
            if min_stock_int < 0:
                messages.error(request, 'El stock minimo no puede ser negativo.')
                errors = True
        except ValueError:
            messages.error(request, 'El stock minimo debe ser un numero entero.')
            errors = True
            min_stock_int = 0

        if errors:
            return render(request, 'productos/form.html', {
                'producto': producto,
                'categorias': categorias,
                'code': code,
                'name': name,
                'description': description,
                'unit': unit,
                'min_stock': min_stock,
                'category_id': category_id,
                'location': location,
            })

        category = get_object_or_404(Category, pk=category_id)

        producto.code = code
        producto.name = name
        producto.description = description or None
        producto.unit = unit
        producto.min_stock = min_stock_int
        producto.category = category
        producto.location = location
        producto.status = status
        producto.save()

        messages.success(request, f'Producto "{name}" actualizado exitosamente.')
        return redirect('producto_list')

    return render(request, 'productos/form.html', {
        'producto': producto,
        'categorias': categorias,
    })


@login_required
def producto_delete(request, pk):
    producto = get_object_or_404(Product, pk=pk)

    entradas_count = Entrada.objects.filter(product=producto).count()
    if entradas_count > 0:
        messages.error(
            request,
            f'No se puede eliminar el producto "{producto.name}" porque tiene {entradas_count} entrada(s) asociada(s).'
        )
        return redirect('producto_list')

    ajustes_count = InventoryAdjustment.objects.filter(product=producto).count()
    if ajustes_count > 0:
        messages.error(
            request,
            f'No se puede eliminar el producto "{producto.name}" porque tiene {ajustes_count} ajuste(s) de inventario asociado(s).'
        )
        return redirect('producto_list')

    if request.method == 'POST':
        name = producto.name
        producto.delete()
        messages.success(request, f'Producto "{name}" eliminado exitosamente.')
        return redirect('producto_list')

    return render(request, 'productos/delete.html', {'producto': producto})


@login_required
def exportar_inventario_actual(request):
    """Exportar reporte de inventario actual a Excel"""
    productos = Product.objects.select_related('category').filter(status='active').order_by('name')

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Actual"

    # Estilos
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    alerta_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    critico_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE INVENTARIO ACTUAL"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fecha de generación
    ws.merge_cells('A2:H2')
    date_cell = ws['A2']
    date_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')

    # Encabezados
    header_row = 4
    headers = ['Codigo', 'Producto', 'Categoria', 'Stock Actual', 'Stock Minimo', 'Unidad', 'Ubicacion', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row_num = header_row + 1
    total_productos = 0
    productos_alerta = 0

    for producto in productos:
        ws.cell(row=row_num, column=1, value=producto.code).border = border
        ws.cell(row=row_num, column=2, value=producto.name).border = border
        ws.cell(row=row_num, column=3, value=producto.category.name).border = border

        stock_cell = ws.cell(row=row_num, column=4, value=producto.stock_actual)
        stock_cell.border = border
        stock_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=row_num, column=5, value=producto.min_stock).border = border
        ws.cell(row=row_num, column=6, value=producto.unit).border = border
        ws.cell(row=row_num, column=7, value=producto.location or 'No especificada').border = border

        # Estado y color según stock
        estado_cell = ws.cell(row=row_num, column=8)
        estado_cell.border = border
        estado_cell.alignment = Alignment(horizontal='center')

        if producto.stock_actual <= 0:
            estado_cell.value = 'SIN STOCK'
            estado_cell.fill = critico_fill
            estado_cell.font = Font(bold=True, color="FFFFFF")
            productos_alerta += 1
        elif producto.stock_actual <= producto.min_stock:
            estado_cell.value = 'STOCK BAJO'
            estado_cell.fill = alerta_fill
            estado_cell.font = Font(bold=True)
            productos_alerta += 1
        else:
            estado_cell.value = 'NORMAL'

        total_productos += 1
        row_num += 1

    # Resumen
    row_num += 2
    ws.merge_cells(f'A{row_num}:C{row_num}')
    resumen_cell = ws.cell(row=row_num, column=1)
    resumen_cell.value = "RESUMEN"
    resumen_cell.font = Font(bold=True, size=12)

    row_num += 1
    ws.cell(row=row_num, column=1, value="Total de productos:")
    ws.cell(row=row_num, column=2, value=total_productos).font = Font(bold=True)

    row_num += 1
    ws.cell(row=row_num, column=1, value="Productos en alerta:")
    ws.cell(row=row_num, column=2, value=productos_alerta).font = Font(bold=True, color="DC3545")

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"inventario_actual_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
