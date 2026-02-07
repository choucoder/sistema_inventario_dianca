from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from ..models import Product, InventarioSesion, DetalleConteo, InventoryAdjustment


@login_required
def inventario_sesiones(request):
    sesiones = InventarioSesion.objects.select_related('user').order_by('-created_at')
    return render(request, 'inventario_fisico/sesiones.html', {'sesiones': sesiones})


@login_required
def inventario_iniciar(request):
    sesion_activa = InventarioSesion.objects.filter(
        user=request.user,
        status='en_proceso'
    ).first()

    if sesion_activa:
        messages.warning(request, 'Ya tiene una sesion de inventario en proceso.')
        return redirect('inventario_conteo', sesion_id=sesion_activa.pk)

    if request.method == 'POST':
        notas = request.POST.get('notas', '').strip()

        sesion = InventarioSesion.objects.create(
            user=request.user,
            notas=notas,
            status='en_proceso'
        )

        messages.success(request, f'Sesion de inventario #{sesion.pk} iniciada. Puede comenzar a escanear productos.')
        return redirect('inventario_conteo', sesion_id=sesion.pk)

    return render(request, 'inventario_fisico/iniciar.html')


@login_required
def inventario_conteo(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.user != request.user and request.user.role != 'admin':
        messages.error(request, 'No tiene permiso para acceder a esta sesion.')
        return redirect('inventario_sesiones')

    if sesion.status in ['finalizado', 'conciliado']:
        return redirect('inventario_resultados', sesion_id=sesion.pk)

    conteos = sesion.detalles.select_related('product').order_by('-created_at')

    return render(request, 'inventario_fisico/conteo.html', {
        'sesion': sesion,
        'conteos': conteos,
    })


@login_required
def inventario_registrar_conteo(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status != 'en_proceso':
        return JsonResponse({
            'success': False,
            'error': 'La sesion no esta en proceso.'
        })

    if request.method == 'POST':
        product_code = request.POST.get('product_code', '').strip()
        cantidad = request.POST.get('cantidad', '').strip()

        try:
            product = Product.objects.get(code__iexact=product_code)
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'No existe producto con codigo "{product_code}"'
            })

        try:
            cantidad_int = int(cantidad)
            if cantidad_int < 0:
                return JsonResponse({
                    'success': False,
                    'error': 'La cantidad no puede ser negativa.'
                })
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'La cantidad debe ser un numero entero.'
            })

        conteo_existente = DetalleConteo.objects.filter(
            sesion=sesion,
            product=product
        ).first()

        stock_sistema = product.stock_actual
        diferencia = cantidad_int - stock_sistema

        if conteo_existente:
            conteo_existente.cantidad_contada = cantidad_int
            conteo_existente.diferencia = diferencia
            conteo_existente.updated_at = timezone.now()
            conteo_existente.save()
            mensaje = f'Conteo actualizado para "{product.name}"'
        else:
            DetalleConteo.objects.create(
                sesion=sesion,
                product=product,
                stock_sistema=stock_sistema,
                cantidad_contada=cantidad_int,
                diferencia=diferencia
            )
            mensaje = f'Conteo registrado para "{product.name}"'

        sesion.total_productos = sesion.detalles.count()
        sesion.save()

        return JsonResponse({
            'success': True,
            'message': mensaje,
            'data': {
                'product_name': product.name,
                'product_code': product.code,
                'stock_sistema': stock_sistema,
                'cantidad_contada': cantidad_int,
                'diferencia': diferencia,
                'unit': product.unit,
            }
        })

    return JsonResponse({'success': False, 'error': 'Metodo no permitido'})


@login_required
def inventario_finalizar(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status != 'en_proceso':
        messages.error(request, 'Esta sesion ya fue finalizada.')
        return redirect('inventario_resultados', sesion_id=sesion.pk)

    if sesion.detalles.count() == 0:
        messages.error(request, 'No puede finalizar una sesion sin conteos.')
        return redirect('inventario_conteo', sesion_id=sesion.pk)

    if request.method == 'POST':
        conteos_con_diferencia = sesion.detalles.exclude(diferencia=0).count()

        sesion.status = 'finalizado'
        sesion.finished_at = timezone.now()
        sesion.productos_con_diferencia = conteos_con_diferencia
        sesion.save()

        if conteos_con_diferencia > 0:
            messages.warning(
                request,
                f'Sesion finalizada. Se encontraron {conteos_con_diferencia} producto(s) con diferencias.'
            )
        else:
            messages.success(
                request,
                'Sesion finalizada. El inventario fisico coincide con el sistema.'
            )

        return redirect('inventario_resultados', sesion_id=sesion.pk)

    return redirect('inventario_conteo', sesion_id=sesion.pk)


@login_required
def inventario_resultados(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status == 'en_proceso':
        return redirect('inventario_conteo', sesion_id=sesion.pk)

    conteos = sesion.detalles.select_related('product').order_by('product__name')
    conteos_con_diferencia = conteos.exclude(diferencia=0)
    conteos_correctos = conteos.filter(diferencia=0)

    return render(request, 'inventario_fisico/resultados.html', {
        'sesion': sesion,
        'conteos': conteos,
        'conteos_con_diferencia': conteos_con_diferencia,
        'conteos_correctos': conteos_correctos,
    })


@login_required
@transaction.atomic
def inventario_conciliar(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status != 'finalizado':
        messages.error(request, 'Solo se pueden conciliar sesiones finalizadas.')
        return redirect('inventario_resultados', sesion_id=sesion.pk)

    if request.method == 'POST':
        conteos_con_diferencia = sesion.detalles.exclude(diferencia=0)

        ajustes_realizados = 0

        for conteo in conteos_con_diferencia:
            InventoryAdjustment.objects.create(
                product=conteo.product,
                user=request.user,
                system_qty=conteo.stock_sistema,
                physical_qty=conteo.cantidad_contada,
                difference=conteo.diferencia
            )

            conteo.product.stock_actual = conteo.cantidad_contada
            conteo.product.save()

            ajustes_realizados += 1

        sesion.status = 'conciliado'
        sesion.conciliated_at = timezone.now()
        sesion.save()

        messages.success(
            request,
            f'Inventario conciliado exitosamente. Se ajustaron {ajustes_realizados} producto(s).'
        )

        return redirect('inventario_resultados', sesion_id=sesion.pk)

    return redirect('inventario_resultados', sesion_id=sesion.pk)


@login_required
def inventario_cancelar(request, sesion_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status not in ['en_proceso', 'finalizado']:
        messages.error(request, 'No se puede cancelar esta sesion.')
        return redirect('inventario_sesiones')

    if request.method == 'POST':
        sesion.status = 'cancelado'
        sesion.save()
        messages.info(request, f'Sesion #{sesion.pk} cancelada.')
        return redirect('inventario_sesiones')

    return render(request, 'inventario_fisico/cancelar.html', {'sesion': sesion})


@login_required
def inventario_eliminar_conteo(request, sesion_id, conteo_id):
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status != 'en_proceso':
        return JsonResponse({
            'success': False,
            'error': 'No se pueden eliminar conteos de sesiones finalizadas.'
        })

    conteo = get_object_or_404(DetalleConteo, pk=conteo_id, sesion=sesion)

    product_name = conteo.product.name
    conteo.delete()

    sesion.total_productos = sesion.detalles.count()
    sesion.save()

    return JsonResponse({
        'success': True,
        'message': f'Conteo de "{product_name}" eliminado.'
    })


@login_required
def exportar_reporte_auditoria(request, sesion_id):
    """Exportar reporte de auditoría de inventario físico a Excel"""
    sesion = get_object_or_404(InventarioSesion, pk=sesion_id)

    if sesion.status == 'en_proceso':
        messages.error(request, 'No se puede generar reporte de una sesion en proceso.')
        return redirect('inventario_conteo', sesion_id=sesion.pk)

    conteos = sesion.detalles.select_related('product', 'product__category').order_by('product__name')

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria de Inventario"

    # Estilos
    header_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    faltante_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    sobrante_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    correcto_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE AUDITORIA - INVENTARIO FISICO"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Info de sesión
    ws.merge_cells('A2:I2')
    info_cell = ws['A2']
    info_cell.value = f"Sesion #{sesion.id} | Iniciada: {sesion.created_at.strftime('%d/%m/%Y %H:%M')} | Estado: {sesion.get_status_display()}"
    info_cell.alignment = Alignment(horizontal='center')

    if sesion.finished_at:
        ws.merge_cells('A3:I3')
        finished_cell = ws['A3']
        finished_cell.value = f"Finalizada: {sesion.finished_at.strftime('%d/%m/%Y %H:%M')} | Usuario: {sesion.user.get_full_name() or sesion.user.username}"
        finished_cell.alignment = Alignment(horizontal='center')
        header_row = 5
    else:
        header_row = 4

    # Encabezados
    headers = ['Codigo', 'Producto', 'Categoria', 'Stock Sistema', 'Stock Fisico', 'Diferencia', 'Unidad', 'Estado', 'Observacion']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row_num = header_row + 1
    productos_correctos = 0
    productos_con_faltante = 0
    productos_con_sobrante = 0

    for conteo in conteos:
        ws.cell(row=row_num, column=1, value=conteo.product.code).border = border
        ws.cell(row=row_num, column=2, value=conteo.product.name).border = border
        ws.cell(row=row_num, column=3, value=conteo.product.category.name).border = border

        sistema_cell = ws.cell(row=row_num, column=4, value=conteo.stock_sistema)
        sistema_cell.border = border
        sistema_cell.alignment = Alignment(horizontal='center')

        fisico_cell = ws.cell(row=row_num, column=5, value=conteo.cantidad_contada)
        fisico_cell.border = border
        fisico_cell.alignment = Alignment(horizontal='center')

        dif_cell = ws.cell(row=row_num, column=6, value=conteo.diferencia)
        dif_cell.border = border
        dif_cell.alignment = Alignment(horizontal='center')
        dif_cell.font = Font(bold=True)

        ws.cell(row=row_num, column=7, value=conteo.product.unit).border = border

        # Estado y observación según diferencia
        estado_cell = ws.cell(row=row_num, column=8)
        estado_cell.border = border
        estado_cell.alignment = Alignment(horizontal='center')

        obs_cell = ws.cell(row=row_num, column=9)
        obs_cell.border = border

        if conteo.diferencia < 0:
            estado_cell.value = 'FALTANTE'
            estado_cell.fill = faltante_fill
            estado_cell.font = Font(bold=True, color="FFFFFF")
            obs_cell.value = f'Faltan {abs(conteo.diferencia)} unidades'
            obs_cell.fill = faltante_fill
            obs_cell.font = Font(color="FFFFFF")
            productos_con_faltante += 1
        elif conteo.diferencia > 0:
            estado_cell.value = 'SOBRANTE'
            estado_cell.fill = sobrante_fill
            estado_cell.font = Font(bold=True, color="FFFFFF")
            obs_cell.value = f'Sobran {conteo.diferencia} unidades'
            obs_cell.fill = sobrante_fill
            obs_cell.font = Font(color="FFFFFF")
            productos_con_sobrante += 1
        else:
            estado_cell.value = 'CORRECTO'
            estado_cell.fill = correcto_fill
            obs_cell.value = 'Inventario coincide'
            obs_cell.fill = correcto_fill
            productos_correctos += 1

        row_num += 1

    # Resumen
    row_num += 2
    ws.merge_cells(f'A{row_num}:D{row_num}')
    resumen_cell = ws.cell(row=row_num, column=1)
    resumen_cell.value = "RESUMEN DE AUDITORIA"
    resumen_cell.font = Font(bold=True, size=14)
    resumen_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    row_num += 1
    ws.cell(row=row_num, column=1, value="Total productos contados:")
    ws.cell(row=row_num, column=2, value=sesion.total_productos).font = Font(bold=True)

    row_num += 1
    ws.cell(row=row_num, column=1, value="Productos correctos:")
    correctos_cell = ws.cell(row=row_num, column=2, value=productos_correctos)
    correctos_cell.font = Font(bold=True, color="28A745")

    row_num += 1
    ws.cell(row=row_num, column=1, value="Productos con faltante:")
    faltante_cell = ws.cell(row=row_num, column=2, value=productos_con_faltante)
    faltante_cell.font = Font(bold=True, color="DC3545")

    row_num += 1
    ws.cell(row=row_num, column=1, value="Productos con sobrante:")
    sobrante_cell = ws.cell(row=row_num, column=2, value=productos_con_sobrante)
    sobrante_cell.font = Font(bold=True, color="28A745")

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"auditoria_inventario_sesion{sesion.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
