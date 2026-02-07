from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from ..models import Salida, Product


@login_required
def salida_registrar(request):
    if request.method == 'POST':
        product_code = request.POST.get('product_code', '').strip()
        receptor = request.POST.get('receptor', '').strip()
        quantity = request.POST.get('quantity', '').strip()
        motivo = request.POST.get('motivo', '').strip()

        errors = False

        if not product_code:
            messages.error(request, 'El codigo del producto es requerido.')
            errors = True
            product = None
        else:
            try:
                product = Product.objects.get(code__iexact=product_code)
                if product.status != 'active':
                    messages.error(request, 'El producto no esta activo.')
                    errors = True
            except Product.DoesNotExist:
                messages.error(request, f'No existe un producto con el codigo "{product_code}".')
                errors = True
                product = None

        if not receptor:
            messages.error(request, 'El receptor es requerido.')
            errors = True

        if not motivo:
            messages.error(request, 'El motivo es requerido.')
            errors = True

        try:
            quantity_int = int(quantity) if quantity else 0
            if quantity_int <= 0:
                messages.error(request, 'La cantidad debe ser mayor a cero.')
                errors = True
        except ValueError:
            messages.error(request, 'La cantidad debe ser un numero entero.')
            errors = True
            quantity_int = 0

        # Validar stock disponible
        if product and quantity_int > 0:
            if product.stock_actual < quantity_int:
                messages.error(
                    request,
                    f'Stock insuficiente. Stock actual: {product.stock_actual} {product.unit}'
                )
                errors = True

        if errors:
            return render(request, 'salidas/registrar.html', {
                'product_code': product_code,
                'receptor': receptor,
                'quantity': quantity,
                'motivo': motivo,
                'product': product if 'product' in dir() and product else None,
            })

        salida = Salida.objects.create(
            product=product,
            user=request.user,
            receptor=receptor,
            quantity=quantity_int,
            motivo=motivo
        )

        stock_msg = ''
        if product.stock_actual <= product.min_stock:
            stock_msg = f' ALERTA: El stock esta en nivel minimo o por debajo ({product.stock_actual}/{product.min_stock}).'

        messages.success(
            request,
            f'Salida registrada exitosamente. Se retiraron {quantity_int} {product.unit} de "{product.name}". Stock restante: {product.stock_actual}.{stock_msg}'
        )
        return redirect('salida_historial')

    return render(request, 'salidas/registrar.html', {
        'product_code': '',
        'receptor': '',
        'quantity': '',
        'motivo': '',
    })


@login_required
def salida_historial(request):
    salidas = Salida.objects.select_related('product', 'user').order_by('-created_at')

    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    producto = request.GET.get('producto', '')
    receptor = request.GET.get('receptor', '')

    if fecha_desde:
        salidas = salidas.filter(created_at__date__gte=fecha_desde)

    if fecha_hasta:
        salidas = salidas.filter(created_at__date__lte=fecha_hasta)

    if producto:
        salidas = salidas.filter(
            Q(product__code__icontains=producto) |
            Q(product__name__icontains=producto)
        )

    if receptor:
        salidas = salidas.filter(receptor__icontains=receptor)

    return render(request, 'salidas/historial.html', {
        'salidas': salidas,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'producto': producto,
            'receptor': receptor,
        }
    })


@login_required
def salida_detalle(request, pk):
    salida = get_object_or_404(
        Salida.objects.select_related('product', 'user'),
        pk=pk
    )
    return render(request, 'salidas/detalle.html', {'salida': salida})


@login_required
def exportar_reporte_salidas(request):
    """Exportar reporte de salidas a Excel"""
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    producto = request.GET.get('producto', '')
    receptor = request.GET.get('receptor', '')

    # Aplicar filtros
    salidas = Salida.objects.select_related('product', 'user').order_by('-created_at')

    if fecha_desde:
        salidas = salidas.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        salidas = salidas.filter(created_at__date__lte=fecha_hasta)
    if producto:
        salidas = salidas.filter(
            Q(product__code__icontains=producto) |
            Q(product__name__icontains=producto)
        )
    if receptor:
        salidas = salidas.filter(receptor__icontains=receptor)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Salidas"

    # Estilos
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE SALIDAS DE PRODUCTOS"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fecha de generación
    ws.merge_cells('A2:H2')
    date_cell = ws['A2']
    date_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')

    # Filtros aplicados
    filtros_texto = []
    if fecha_desde:
        filtros_texto.append(f"Desde: {fecha_desde}")
    if fecha_hasta:
        filtros_texto.append(f"Hasta: {fecha_hasta}")
    if producto:
        filtros_texto.append(f"Producto: {producto}")
    if receptor:
        filtros_texto.append(f"Receptor: {receptor}")

    if filtros_texto:
        ws.merge_cells('A3:H3')
        filtro_cell = ws['A3']
        filtro_cell.value = "Filtros: " + " | ".join(filtros_texto)
        filtro_cell.alignment = Alignment(horizontal='center')
        header_row = 5
    else:
        header_row = 4

    # Encabezados
    headers = ['Fecha', 'Codigo', 'Producto', 'Cantidad', 'Unidad', 'Receptor', 'Registrado por', 'Motivo']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row_num = header_row + 1
    for salida in salidas:
        ws.cell(row=row_num, column=1, value=salida.created_at.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row_num, column=2, value=salida.product.code).border = border
        ws.cell(row=row_num, column=3, value=salida.product.name).border = border
        ws.cell(row=row_num, column=4, value=salida.quantity).border = border
        ws.cell(row=row_num, column=5, value=salida.product.unit).border = border
        ws.cell(row=row_num, column=6, value=salida.receptor).border = border
        ws.cell(row=row_num, column=7, value=salida.user.get_full_name() or salida.user.username).border = border
        ws.cell(row=row_num, column=8, value=salida.motivo).border = border
        row_num += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 40

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_salidas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
